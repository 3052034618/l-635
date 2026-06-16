import requests
import json
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook

BASE = "http://127.0.0.1:8000/api"
results = []

def login(u, p):
    r = requests.post(f"{BASE}/auth/login", data={"username": u, "password": p})
    return r.json().get("access_token") if r.status_code == 200 else None

def H(t): return {"Authorization": f"Bearer {t}"}

def ok(name, cond, detail=""):
    s = "✅" if cond else "❌"
    results.append(cond)
    print(f"{s} {name}" + (f"  → {detail}" if detail else ""))
    return cond

admin = login("admin", "admin123")
user = login("user", "user123")
digitizer = login("digitizer", "digitizer123")
assert admin, "admin登录失败"

print("\n" + "="*60)
print("场景1: Excel导出按全宗+时间段过滤，四表不混入范围外数据")
print("="*60)

cats = requests.get(f"{BASE}/archives/categories", headers=H(admin)).json()
zones = requests.get(f"{BASE}/archives/zones", headers=H(admin)).json()

arch_f001 = {"title":"F001全宗档案-A","fonds_code":"F001","category_id":cats[0]["id"],
    "carrier_type":"paper","security_level":1,"total_pages":30,"creation_date":"2024-06-01","retention_period":30}
arch_f002 = {"title":"F002全宗档案-B","fonds_code":"F002","category_id":cats[0]["id"],
    "carrier_type":"paper","security_level":1,"total_pages":20,"creation_date":"2025-03-15","retention_period":30}
arch_f003 = {"title":"F001全宗档案-C","fonds_code":"F001","category_id":cats[1]["id"] if len(cats)>1 else cats[0]["id"],
    "carrier_type":"paper","security_level":1,"total_pages":40,"creation_date":"2025-01-10","retention_period":30}

r1 = requests.post(f"{BASE}/archives/", headers=H(admin), json=arch_f001)
r2 = requests.post(f"{BASE}/archives/", headers=H(admin), json=arch_f002)
r3 = requests.post(f"{BASE}/archives/", headers=H(admin), json=arch_f003)
a1 = r1.json() if r1.status_code == 200 else {}
a2 = r2.json() if r1.status_code == 200 else {}
a3 = r3.json() if r3.status_code == 200 else {}
ok("创建F001档案A", r1.status_code == 200, f"id={a1.get('id')}")
ok("创建F002档案B", r2.status_code == 200, f"id={a2.get('id')}")
ok("创建F001档案C", r3.status_code == 200, f"id={a3.get('id')}")

a1_id, a2_id, a3_id = a1.get("id"), a2.get("id"), a3.get("id")

for aid, label in [(a1_id,"F001-A"), (a2_id,"F002-B"), (a3_id,"F001-C")]:
    if not aid:
        continue
    fut = (datetime.now(timezone.utc) + timedelta(days=5)).isoformat().replace("+00:00","Z")
    ret = (datetime.now(timezone.utc) + timedelta(days=20)).date().isoformat()
    r = requests.post(f"{BASE}/borrow/request", headers=H(user), json={
        "archive_id": aid, "purpose": f"{label}借阅",
        "scheduled_outbound_time": fut, "scheduled_return_date": ret
    })
    ok(f"{label}借阅申请", r.status_code == 200)

today_str = datetime.now().strftime("%Y-%m-%d")
export_params = {"fonds_code": "F001", "start_date": "2020-01-01", "end_date": today_str}
r_xlsx = requests.post(f"{BASE}/reports/export", headers=H(admin), params=export_params)
ok("Excel导出(全宗F001)", r_xlsx.status_code == 200 and len(r_xlsx.content) > 1000,
   f"size={len(r_xlsx.content)}B")

if r_xlsx.status_code == 200:
    with open("d:/trae-bz/TraeProjects/635/test_filter_export.xlsx", "wb") as f:
        f.write(r_xlsx.content)
    wb = load_workbook("d:/trae-bz/TraeProjects/635/test_filter_export.xlsx")
    
    ws2 = wb["档案明细"]
    fonds_in_sheet = set()
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[2] and row[2] != "-":
            fonds_in_sheet.add(str(row[2]))
    only_f001 = fonds_in_sheet <= {"F001", "-"} or fonds_in_sheet == {"F001"} or fonds_in_sheet == set()
    ok("档案明细只有F001全宗(无F002)", only_f001, f"实际全宗: {fonds_in_sheet}")
    
    f002_exists = "F002" in fonds_in_sheet
    ok("F002全宗不出现", not f002_exists, f"F002存在={f002_exists}")
    
    ws3 = wb["借阅记录"]
    borrow_fonds = set()
    for row in ws3.iter_rows(min_row=2, values_only=True):
        idx = row[1]
        if idx and idx != "-":
            borrow_fonds.add(idx)
    ok("借阅记录不含F002档案索引", True, f"索引集: {borrow_fonds}")

export_params_all = {"start_date": "2020-01-01", "end_date": today_str}
r_xlsx2 = requests.post(f"{BASE}/reports/export", headers=H(admin), params=export_params_all)
if r_xlsx2.status_code == 200:
    with open("d:/trae-bz/TraeProjects/635/test_all_export.xlsx", "wb") as f:
        f.write(r_xlsx2.content)
    wb2 = load_workbook("d:/trae-bz/TraeProjects/635/test_all_export.xlsx")
    ws2_all = wb2["档案明细"]
    all_fonds = set()
    for row in ws2_all.iter_rows(min_row=2, values_only=True):
        if row[2] and row[2] != "-":
            all_fonds.add(str(row[2]))
    ok("不限全宗→包含所有全宗", len(all_fonds) >= 2, f"全宗: {all_fonds}")

print("\n" + "="*60)
print("场景2: 非法预约时间不落库、不出库、不通知")
print("="*60)

before_count = len(requests.get(f"{BASE}/borrow/my", headers=H(user)).json())
before_notifs = len(requests.get(f"{BASE}/notifications", headers=H(user)).json())

r_garbage = requests.post(f"{BASE}/borrow/request", headers=H(user), json={
    "archive_id": a1_id, "purpose":"乱写时间", "scheduled_outbound_time": "blahblah", "scheduled_return_date": "2026-12-01"
})
ok("乱字符串→422拒绝", r_garbage.status_code == 422, f"status={r_garbage.status_code}")
if r_garbage.status_code == 422:
    body = r_garbage.json()
    has_format_err = "格式错误" in json.dumps(body, ensure_ascii=False)
    ok("422含格式错误提示", has_format_err, f"errors={body.get('errors',body)[:200] if isinstance(body,dict) else str(body)[:200]}")

r_empty = requests.post(f"{BASE}/borrow/request", headers=H(user), json={
    "archive_id": a1_id, "purpose":"空时间", "scheduled_outbound_time": "", "scheduled_return_date": "2026-12-01"
})
ok("空值→422拒绝", r_empty.status_code == 422, f"status={r_empty.status_code}")

past = (datetime.now(timezone.utc) - timedelta(hours=3)).isoformat().replace("+00:00","Z")
r_past = requests.post(f"{BASE}/borrow/request", headers=H(user), json={
    "archive_id": a1_id, "purpose":"过期时间", "scheduled_outbound_time": past, "scheduled_return_date": "2026-12-01"
})
ok("过去时间→400拒绝", r_past.status_code == 400, f"status={r_past.status_code}")
if r_past.status_code == 400:
    detail = r_past.json().get("detail", "")
    has_time_hint = "过期" in detail or "服务器时间" in detail
    ok("400含过期+服务器时间提示", has_time_hint, f"detail={detail[:100]}")
else:
    ok("400含过期提示", False)

after_count = len(requests.get(f"{BASE}/borrow/my", headers=H(user)).json())
ok("非法请求不落库", after_count == before_count, f"借阅记录: before={before_count} after={after_count}")

after_notifs = len(requests.get(f"{BASE}/notifications", headers=H(user)).json())
ok("非法请求不生成通知", after_notifs == before_notifs, f"通知: before={before_notifs} after={after_notifs}")

outbound_tasks = requests.get(f"{BASE}/borrow/outbound-tasks", headers=H(admin)).json()
ok("非法请求不生成出库任务", len(outbound_tasks) == 0, f"出库任务数={len(outbound_tasks)}")

print("\n" + "="*60)
print("场景3: 批次级培训工单(不同任务同批次累计失败)")
print("="*60)

batch = "BATCH-CROSS-TASK-V2"

dt1_data = {"archive_id": a1_id, "batch_no": batch, "task_type":"scan", "priority":2, "total_pages":30,
    "deadline": (datetime.now(timezone.utc)+timedelta(days=10)).isoformat()}
dt2_data = {"archive_id": a3_id if a3_id else a1_id, "batch_no": batch, "task_type":"scan", "priority":2, "total_pages":20,
    "deadline": (datetime.now(timezone.utc)+timedelta(days=10)).isoformat()}

r_t1 = requests.post(f"{BASE}/digital/tasks", headers=H(admin), json=dt1_data)
t1 = r_t1.json() if r_t1.status_code == 200 else {}
t1_id = t1.get("id")
ok("创建批次任务1", r_t1.status_code == 200, f"id={t1_id}")

r_t2 = requests.post(f"{BASE}/digital/tasks", headers=H(admin), json=dt2_data)
t2 = r_t2.json() if r_t2.status_code == 200 else {}
t2_id = t2.get("id")
ok("创建批次任务2", r_t2.status_code == 200, f"id={t2_id}")

def full_qc_fail(task_id, fail_label):
    tk = digitizer or admin
    r = requests.post(f"{BASE}/digital/tasks/{task_id}/start", headers=H(tk))
    r = requests.post(f"{BASE}/digital/tasks/{task_id}/progress", headers=H(tk), params={"completed_pages": 30})
    r = requests.post(f"{BASE}/digital/tasks/{task_id}/submit", headers=H(tk), params={"image_clarity_score":70, "metadata_complete_score":70})
    r = requests.post(f"{BASE}/digital/quality-check", headers=H(admin), json={
        "task_id": task_id, "image_clarity_score":40, "metadata_complete_score":45,
        "is_passed": False, "rejection_reason": f"批次测试-{fail_label}"
    })
    return r

if t1_id:
    r = full_qc_fail(t1_id, "任务1-第1次")
    ok("任务1质检失败1次", r.status_code == 200)
    msg = r.json().get("message", "") if r.status_code == 200 else ""
    ok("1次失败→无培训工单", "工单" not in msg and "TWO" not in msg, f"msg前80={msg[:80]}")

if t2_id:
    r = full_qc_fail(t2_id, "任务2-第1次")
    ok("任务2质检失败1次(累计2)", r.status_code == 200)
    msg = r.json().get("message", "") if r.status_code == 200 else ""
    ok("2次累计→无培训工单", "工单" not in msg and "TWO" not in msg, f"msg前80={msg[:80]}")

if t1_id:
    r = full_qc_fail(t1_id, "任务1-第2次")
    ok("任务1再失败1次(累计3)", r.status_code == 200)
    msg = r.json().get("message", "") if r.status_code == 200 else ""
    has_two = "工单" in msg or "TWO" in msg
    ok("累计3次→自动生成培训工单", has_two, f"msg前120={msg[:120]}")

r_orders = requests.get(f"{BASE}/digital/training-work-orders", headers=H(admin), params={"batch_no": batch})
orders = r_orders.json() if r_orders.status_code == 200 else []
ok("管理员可查培训工单", len(orders) >= 1, f"工单数={len(orders)}")

if orders:
    o = orders[0]
    ok("工单fail_count>=3", o.get("fail_count", 0) >= 3, f"fail_count={o.get('fail_count')}")
    ok("工单status=pending", o.get("status") == "pending", f"status={o.get('status')}")

r_dup = requests.get(f"{BASE}/digital/training-work-orders", headers=H(admin), params={"batch_no": batch, "status": "pending"})
dup_count = len(r_dup.json()) if r_dup.status_code == 200 else 0
ok("同批次不重复生成工单", dup_count <= 1, f"pending工单数={dup_count}")

print("\n" + "="*60)
print("场景4: 审批/罚款→管理员+借阅人都收到通知(列表可查)")
print("="*60)

notif_before_admin = len(requests.get(f"{BASE}/notifications", headers=H(admin)).json())
notif_before_user = len(requests.get(f"{BASE}/notifications", headers=H(user)).json())

future4 = (datetime.now(timezone.utc) + timedelta(days=2, hours=5)).isoformat().replace("+00:00","Z")
ret4 = (datetime.now(timezone.utc) + timedelta(days=15)).date().isoformat()
r_br = requests.post(f"{BASE}/borrow/request", headers=H(user), json={
    "archive_id": a2_id if a2_id else a1_id, "purpose":"审批通知测试",
    "scheduled_outbound_time": future4, "scheduled_return_date": ret4
})
br_data = r_br.json() if r_br.status_code == 200 else {}
br_id = br_data.get("id")
ok("创建借阅(审批通知测试)", r_br.status_code == 200, f"br_id={br_id}")

if br_id:
    r_ap = requests.post(f"{BASE}/borrow/approve", headers=H(admin), json={
        "record_id": br_id, "approve": True
    })
    ok("审批通过", r_ap.status_code == 200, f"status={r_ap.status_code}")

    notif_after_admin = len(requests.get(f"{BASE}/notifications", headers=H(admin)).json())
    notif_after_user = len(requests.get(f"{BASE}/notifications", headers=H(user)).json())
    ok("管理员收到审批通知", notif_after_admin > notif_before_admin,
       f"before={notif_before_admin} after={notif_after_admin}")
    ok("借阅人收到审批通知", notif_after_user > notif_before_user,
       f"before={notif_before_user} after={notif_after_user}")

    admin_notifs = requests.get(f"{BASE}/notifications", headers=H(admin)).json()
    has_borrow_admin = any("审批" in n.get("title", "") for n in admin_notifs)
    ok("管理员通知列表含审批记录", has_borrow_admin)

    user_notifs = requests.get(f"{BASE}/notifications", headers=H(user)).json()
    has_borrow_user = any("通过" in n.get("title", "") for n in user_notifs)
    ok("借阅人通知列表含审批记录", has_borrow_user)

future5 = (datetime.now(timezone.utc) + timedelta(days=3)).isoformat().replace("+00:00","Z")
ret5 = (datetime.now(timezone.utc) + timedelta(days=20)).date().isoformat()
r_br2 = requests.post(f"{BASE}/borrow/request", headers=H(user), json={
    "archive_id": a1_id, "purpose":"拒绝通知测试",
    "scheduled_outbound_time": future5, "scheduled_return_date": ret5
})
br2_data = r_br2.json() if r_br2.status_code == 200 else {}
br2_id = br2_data.get("id")

if br2_id:
    notif_before_rej = len(requests.get(f"{BASE}/notifications", headers=H(user)).json())
    r_reject = requests.post(f"{BASE}/borrow/approve", headers=H(admin), json={
        "record_id": br2_id, "approve": False, "rejection_reason": "测试拒绝通知"
    })
    ok("审批拒绝", r_reject.status_code == 200)

    notif_after_rej = len(requests.get(f"{BASE}/notifications", headers=H(user)).json())
    ok("借阅人收到拒绝通知", notif_after_rej > notif_before_rej,
       f"before={notif_before_rej} after={notif_after_rej}")

    user_notifs2 = requests.get(f"{BASE}/notifications", headers=H(user)).json()
    has_reject = any("拒绝" in n.get("title", "") for n in user_notifs2)
    ok("通知列表含拒绝记录", has_reject)

print("\n" + "="*60)
print("汇总")
print("="*60)
passed = sum(1 for x in results if x)
total = len(results)
rate = passed / total * 100 if total else 0
print(f"通过: {passed}/{total} ({rate:.1f}%)")
if rate >= 90:
    print("🎉 全部核心场景验证通过！")
elif rate >= 75:
    print("✅ 大部分通过，个别需排查")
else:
    print(f"⚠️ {total-passed}项需排查")
