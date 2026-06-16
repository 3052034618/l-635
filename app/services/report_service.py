from typing import Optional, List
from datetime import datetime, date
from io import BytesIO
from sqlalchemy.orm import Session
from sqlalchemy import and_

from app.models.report import MonthlyReport
from app.models.archive import Archive
from app.models.archive_category import ArchiveCategory
from app.models.borrow import BorrowRecord
from app.models.digital import DigitalTask
from app.models.monitoring import SensorReading, WorkOrder, Sensor
from app.models.storage import StorageZone, StorageCabinet
from app.models.user import User
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


class ReportService:
    @staticmethod
    def _is_date_in_range(
        dt_value: Optional[datetime],
        start_date: Optional[date],
        end_date: Optional[date]
    ) -> bool:
        if not dt_value:
            return False
        d = dt_value.date() if isinstance(dt_value, datetime) else dt_value
        if start_date and d < start_date:
            return False
        if end_date and d > end_date:
            return False
        return True

    @staticmethod
    def _in_month(dt_value: Optional[datetime], year: int, month: int) -> bool:
        if not dt_value:
            return False
        d = dt_value.date() if isinstance(dt_value, datetime) else dt_value
        return d.year == year and d.month == month

    @staticmethod
    def generate_monthly_report(
        db: Session,
        report_month: Optional[str] = None
    ) -> List[MonthlyReport]:
        if report_month is None:
            today = date.today()
            year, month = today.year, today.month
        else:
            parts = report_month.split("-")
            year, month = int(parts[0]), int(parts[1])
        report_month_str = f"{year}-{str(month).zfill(2)}"
        
        zones = db.query(StorageZone).all()
        if not zones:
            return []
        
        reports = []
        
        for zone in zones:
            cabinets = db.query(StorageCabinet).filter(StorageCabinet.zone_id == zone.id).all()
            cabinet_ids = [c.id for c in cabinets]
            
            new_archives_count = 0
            all_archives = db.query(Archive).filter(Archive.cabinet_id.in_(cabinet_ids)).all() if cabinet_ids else []
            for a in all_archives:
                if ReportService._in_month(a.created_at, year, month):
                    new_archives_count += 1
            
            borrow_count = 0
            if cabinet_ids:
                archive_ids_subq = [a.id for a in all_archives]
                all_borrows = db.query(BorrowRecord).filter(
                    BorrowRecord.archive_id.in_(archive_ids_subq)
                ).all() if archive_ids_subq else []
                for b in all_borrows:
                    if ReportService._in_month(b.borrow_date, year, month):
                        borrow_count += 1
            
            digitization_count = 0
            if cabinet_ids:
                all_digital_tasks = db.query(DigitalTask).filter(
                    DigitalTask.status == "completed",
                    DigitalTask.archive_id.in_([a.id for a in all_archives])
                ).all() if all_archives else []
                for t in all_digital_tasks:
                    if ReportService._in_month(t.completed_at, year, month):
                        digitization_count += 1
            
            total_archives_in_zone = len(all_archives)
            digitized_count = sum(1 for a in all_archives if a.is_digitized)
            digitization_rate = round((digitized_count / total_archives_in_zone * 100), 2) if total_archives_in_zone > 0 else 0.0
            
            sensors = db.query(Sensor).filter(Sensor.zone_id == zone.id).all()
            sensor_ids = [s.id for s in sensors]
            
            temp_warning_count = 0
            humidity_warning_count = 0
            if sensor_ids:
                readings = db.query(SensorReading).filter(SensorReading.sensor_id.in_(sensor_ids)).all()
                for r in readings:
                    if not ReportService._in_month(r.reading_time, year, month):
                        continue
                    if r.is_warning:
                        if r.temperature is not None:
                            if r.temperature < zone.temperature_min or r.temperature > zone.temperature_max:
                                temp_warning_count += 1
                        if r.humidity is not None:
                            if r.humidity < zone.humidity_min or r.humidity > zone.humidity_max:
                                humidity_warning_count += 1
            
            all_work_orders = db.query(WorkOrder).filter(WorkOrder.zone_id == zone.id).all()
            total_warning_count = sum(1 for w in all_work_orders if ReportService._in_month(w.created_at, year, month))
            
            existing = db.query(MonthlyReport).filter(
                MonthlyReport.report_month == report_month_str,
                MonthlyReport.zone_code == zone.code
            ).first()
            
            if existing:
                existing.new_archives_count = new_archives_count
                existing.borrow_count = borrow_count
                existing.digitization_count = digitization_count
                existing.digitization_rate = digitization_rate
                existing.temp_warning_count = temp_warning_count
                existing.humidity_warning_count = humidity_warning_count
                existing.total_warning_count = total_warning_count
                existing.generated_at = datetime.utcnow()
                report = existing
            else:
                report = MonthlyReport(
                    report_month=report_month_str,
                    zone_code=zone.code,
                    new_archives_count=new_archives_count,
                    borrow_count=borrow_count,
                    digitization_count=digitization_count,
                    digitization_rate=digitization_rate,
                    temp_warning_count=temp_warning_count,
                    humidity_warning_count=humidity_warning_count,
                    total_warning_count=total_warning_count
                )
                db.add(report)
            
            reports.append(report)
        
        db.commit()
        for r in reports:
            db.refresh(r)
        return reports

    @staticmethod
    def get_monthly_reports(
        db: Session,
        report_month: Optional[str] = None,
        zone_code: Optional[str] = None
    ) -> List[MonthlyReport]:
        query = db.query(MonthlyReport)
        if report_month:
            query = query.filter(MonthlyReport.report_month == report_month)
        if zone_code:
            query = query.filter(MonthlyReport.zone_code == zone_code)
        return query.order_by(MonthlyReport.report_month.desc(), MonthlyReport.zone_code).all()

    @staticmethod
    def export_report_to_excel(
        db: Session,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        zone_code: Optional[str] = None,
        fonds_code: Optional[str] = None
    ) -> BytesIO:
        wb = Workbook()
        header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
        header_font = Font(bold=True, color="FF1F3864")
        center = Alignment(horizontal="center", vertical="center")

        all_cabinets = db.query(StorageCabinet).all()
        cabinet_map = {c.id: c for c in all_cabinets}
        all_zones = db.query(StorageZone).all()
        zones_map = {z.id: z for z in all_zones}
        all_categories = db.query(ArchiveCategory).all()
        cat_map = {c.id: c for c in all_categories}
        all_users = db.query(User).all()
        user_map = {u.id: u for u in all_users}
        all_sensors = db.query(Sensor).all()
        sensor_map = {s.id: s for s in all_sensors}

        zones_query = db.query(StorageZone)
        if zone_code:
            zones_query = zones_query.filter(StorageZone.code == zone_code)
        filtered_zones = zones_query.all()

        archive_query = db.query(Archive)
        if fonds_code:
            archive_query = archive_query.filter(Archive.fonds_code == fonds_code)
        all_archives_raw = archive_query.all()
        filtered_archives = []
        for a in all_archives_raw:
            if not ReportService._is_date_in_range(a.created_at, start_date, end_date):
                continue
            if zone_code:
                cab = cabinet_map.get(a.cabinet_id)
                if cab:
                    z = zones_map.get(cab.zone_id)
                    if not z or z.code != zone_code:
                        continue
            filtered_archives.append(a)
        filtered_archive_ids = {a.id for a in filtered_archives}
        archive_map = {a.id: a for a in all_archives_raw}

        all_borrows_raw = db.query(BorrowRecord).all()
        filtered_borrows = []
        for b in all_borrows_raw:
            if b.archive_id not in filtered_archive_ids:
                continue
            if not ReportService._is_date_in_range(b.borrow_date, start_date, end_date):
                continue
            filtered_borrows.append(b)

        zone_filtered_cabinet_ids = set()
        for z in filtered_zones:
            for c in all_cabinets:
                if c.zone_id == z.id:
                    zone_filtered_cabinet_ids.add(c.id)
        zone_filtered_sensor_ids = set()
        for s in all_sensors:
            if s.zone_id in {z.id for z in filtered_zones}:
                zone_filtered_sensor_ids.add(s.id)

        ws1 = wb.active
        ws1.title = "运营总览"
        headers1 = ["库区代码", "库区名称", "新增归档量", "借阅次数", "数字化数量", "数字化转化率(%)", "温度超标次数", "湿度超标次数", "总预警次数"]
        ws1.append(headers1)
        for col in range(1, len(headers1) + 1):
            cell = ws1.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for zone in filtered_zones:
            zone_cabinet_ids = {c.id for c in all_cabinets if c.zone_id == zone.id}
            zone_archives = [a for a in filtered_archives if a.cabinet_id in zone_cabinet_ids]
            new_count = len(zone_archives)
            zone_archive_ids = {a.id for a in zone_archives}
            zone_borrows = [b for b in filtered_borrows if b.archive_id in zone_archive_ids]
            borrow_ct = len(zone_borrows)
            digi_count = sum(1 for a in zone_archives if a.is_digitized)
            digi_rate = round((digi_count / new_count * 100), 2) if new_count > 0 else 0.0
            zone_sensors = [s for s in all_sensors if s.zone_id == zone.id]
            zone_sensor_ids = {s.id for s in zone_sensors}
            zone_readings = db.query(SensorReading).filter(
                SensorReading.sensor_id.in_(zone_sensor_ids)
            ).all() if zone_sensor_ids else []
            temp_warn = 0
            humidity_warn = 0
            for r in zone_readings:
                if not ReportService._is_date_in_range(r.reading_time, start_date, end_date):
                    continue
                if r.is_warning:
                    if r.temperature is not None and (r.temperature < zone.temperature_min or r.temperature > zone.temperature_max):
                        temp_warn += 1
                    if r.humidity is not None and (r.humidity < zone.humidity_min or r.humidity > zone.humidity_max):
                        humidity_warn += 1
            zone_work_orders = db.query(WorkOrder).filter(WorkOrder.zone_id == zone.id).all()
            total_warn = sum(1 for w in zone_work_orders if ReportService._is_date_in_range(w.created_at, start_date, end_date))
            ws1.append([
                zone.code, zone.name,
                new_count, borrow_ct,
                digi_count, digi_rate,
                temp_warn, humidity_warn,
                total_warn
            ])

        for col in ws1.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws1.column_dimensions[column].width = min(max_length + 4, 30)

        ws2 = wb.create_sheet("档案明细")
        headers2 = ["索引号", "题名", "全宗号", "分类", "载体类型", "密级", "是否数字化", "数字化质量(%)", "入库日期", "状态", "所在库区", "所在柜位"]
        ws2.append(headers2)
        for col in range(1, len(headers2) + 1):
            cell = ws2.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for a in filtered_archives:
            cat = cat_map.get(a.category_id)
            cab = cabinet_map.get(a.cabinet_id)
            zone = zones_map.get(cab.zone_id) if cab else None
            ws2.append([
                a.archive_index, a.title, a.fonds_code or "-",
                cat.name if cat else "-", a.carrier_type, a.security_level,
                "是" if a.is_digitized else "否",
                round(a.digitization_quality, 2),
                str(a.storage_start_date) if a.storage_start_date else "-",
                a.status, zone.name if zone else "-",
                a.cabinet_slot or "-"
            ])

        for col in ws2.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws2.column_dimensions[column].width = min(max_length + 4, 50)

        ws3 = wb.create_sheet("借阅记录")
        headers3 = ["借阅编号", "档案索引号", "用户ID", "用户部门", "预约出库时间", "应还日期", "实还日期", "状态", "审批状态", "逾期天数", "罚款金额(元)"]
        ws3.append(headers3)
        for col in range(1, len(headers3) + 1):
            cell = ws3.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for br in filtered_borrows:
            archive = archive_map.get(br.archive_id)
            user = user_map.get(br.user_id)
            ws3.append([
                br.record_no,
                archive.archive_index if archive else "-",
                br.user_id,
                user.department if user else "-",
                br.scheduled_outbound_time.strftime("%Y-%m-%d %H:%M") if br.scheduled_outbound_time else "-",
                str(br.scheduled_return_date),
                str(br.actual_return_date) if br.actual_return_date else "-",
                br.status, br.approval_status,
                br.overdue_days, br.fine_amount
            ])

        for col in ws3.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws3.column_dimensions[column].width = min(max_length + 4, 30)

        ws4 = wb.create_sheet("温湿度记录")
        headers4 = ["记录时间", "库区", "传感器", "温度(°C)", "湿度(%)", "是否超标预警"]
        ws4.append(headers4)
        for col in range(1, len(headers4) + 1):
            cell = ws4.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        readings_query = db.query(SensorReading).order_by(SensorReading.reading_time.desc()).limit(2000)
        if zone_filtered_sensor_ids:
            readings_query = readings_query.filter(SensorReading.sensor_id.in_(zone_filtered_sensor_ids))
        all_readings = readings_query.all()

        filtered_readings = []
        for r in all_readings:
            if not ReportService._is_date_in_range(r.reading_time, start_date, end_date):
                continue
            filtered_readings.append(r)

        for r in filtered_readings[:500]:
            sensor = sensor_map.get(r.sensor_id)
            zone = zones_map.get(sensor.zone_id) if sensor else None
            ws4.append([
                r.reading_time.strftime("%Y-%m-%d %H:%M:%S"),
                zone.name if zone else "-",
                sensor.name if sensor else "-",
                round(r.temperature, 2) if r.temperature is not None else "-",
                round(r.humidity, 2) if r.humidity is not None else "-",
                "是" if r.is_warning else "否"
            ])

        for col in ws4.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws4.column_dimensions[column].width = min(max_length + 4, 25)

        info_row = 30
        ws1.cell(row=info_row, column=1, value="导出时间:").font = Font(bold=True)
        ws1.cell(row=info_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws1.cell(row=info_row + 1, column=1, value="统计范围:").font = Font(bold=True)
        ws1.cell(row=info_row + 1, column=2, value=f"{start_date or '不限'} 至 {end_date or '不限'}")
        ws1.cell(row=info_row + 2, column=1, value="全宗号筛选:").font = Font(bold=True)
        ws1.cell(row=info_row + 2, column=2, value=fonds_code or "全部")
        ws1.cell(row=info_row + 3, column=1, value="库区筛选:").font = Font(bold=True)
        ws1.cell(row=info_row + 3, column=2, value=zone_code or "全部")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def add_operation_log(
        db: Session,
        user_id: int,
        operation_type: str,
        target_type: str,
        target_id: int,
        description: str,
        ip_address: Optional[str] = None
    ):
        from app.models.report import OperationLog
        log = OperationLog(
            user_id=user_id,
            operation_type=operation_type,
            target_type=target_type,
            target_id=target_id,
            description=description,
            ip_address=ip_address
        )
        db.add(log)
        db.commit()
